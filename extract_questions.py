#!/usr/bin/env python3
"""
Extract questions from HTML file and save to Excel
"""

from bs4 import BeautifulSoup
import pandas as pd
import re
import sys
import os

def clean_text(text):
    """Clean HTML text by removing extra spaces and special characters"""
    if not text:
        return ""
    
    # Remove extra whitespace
    text = re.sub(r'\s+', ' ', text)
    
    # Remove special characters like Â
    text = text.replace('Â', '')
    text = text.replace('\xa0', ' ')
    
    # Clean up common HTML entities
    text = text.replace('&nbsp;', ' ')
    text = text.replace('&amp;', '&')
    text = text.replace('&lt;', '<')
    text = text.replace('&gt;', '>')
    
    # Remove leading/trailing whitespace
    text = text.strip()
    
    return text

def extract_formula_from_element(element):
    """Extract mathematical formulas from HTML element including images"""
    if not element:
        return ""
    
    # Convert element to string to process
    html_content = str(element)
    
    # Find all img tags with formulas
    soup = BeautifulSoup(html_content, 'html.parser')
    
    # Replace img tags with their URLs
    for img in soup.find_all('img'):
        src = img.get('src', '')
        if 'cloudfront.net' in src and src.endswith('.png'):
            # Just put the URL directly
            img.replace_with(f" {src} ")
    
    # Get the modified text
    text = soup.get_text()
    
    # Clean the text
    return clean_text(text)

def extract_answer_key(soup):
    """Extract answer key from HTML"""
    answer_key = {}
    
    # Find the Answer Key section - handle whitespace in header
    answer_key_header = None
    all_h3 = soup.find_all('h3')
    
    for h3 in all_h3:
        h3_text = h3.get_text().strip()
        if 'Answer Key' in h3_text:
            answer_key_header = h3
            break
    
    if not answer_key_header:
        return answer_key
    
    # Start from the header and find all question answers
    current = answer_key_header.find_next_sibling()
    
    count = 0
    while current and count < 200:  # Process up to 200 questions
        if current.name == 'p':
            classes = current.get('class', [])
            
            if 'pull-left' in classes and 'clearfix' in classes:
                # Get the raw text including all whitespace
                raw_text = current.get_text()
                
                # Extract question number
                q_match = re.search(r'Q\.(\d+)\)', raw_text)
                if q_match:
                    question_num = int(q_match.group(1))
                    
                    # Simply find any letter A-E in the text after Q.X)
                    # This is more robust than trying to parse exact positions
                    for match in re.finditer(r'[A-E]', raw_text):
                        char = match.group()
                        # Make sure it's after the Q.X) part
                        if match.start() > q_match.end():
                            answer_key[question_num] = char
                            break
                    
                    count += 1
        
        # Move to next sibling
        current = current.find_next_sibling()
        
        # Stop if we hit another section header
        if current and current.name == 'h3':
            break
    
    return answer_key

def extract_explanations(soup):
    """Extract explanations from HTML"""
    explanations = {}
    
    # Find the Explanation section - handle whitespace
    explanation_header = None
    all_h3 = soup.find_all('h3')
    
    for h3 in all_h3:
        if 'Explanation' in h3.get_text().strip():
            explanation_header = h3
            break
    
    if not explanation_header:
        return explanations
    
    # Start from the header and find all explanations
    current = explanation_header.find_next_sibling()
    
    count = 0
    while current and count < 200:
        if current.name == 'p':
            classes = current.get('class', [])
            # Check for pull-left class (without requiring clearfix)
            if 'pull-left' in classes:
                # Get the full content
                raw_text = current.get_text()
                
                # Extract question number
                q_match = re.match(r'Q\.(\d+)\)', raw_text.strip())
                if q_match:
                    question_num = int(q_match.group(1))
                    
                    # Find ALL explanation text in nested p tags
                    explanation_paragraphs = current.find_all('p')
                    if explanation_paragraphs:
                        # Combine all nested paragraph texts with formula extraction
                        explanation_parts = []
                        for p in explanation_paragraphs:
                            # Use formula extraction for each paragraph
                            text = extract_formula_from_element(p)
                            if text:
                                explanation_parts.append(text)
                        
                        # Join all parts with a space
                        full_explanation = ' '.join(explanation_parts)
                        if full_explanation:
                            explanations[question_num] = full_explanation
                    else:
                        # If no nested p tags, extract with formulas
                        full_text = extract_formula_from_element(current)
                        # Remove the Q.X) part
                        explanation_text = re.sub(r'^Q\.\d+\)\s*', '', full_text)
                        if explanation_text:
                            explanations[question_num] = explanation_text
                    
                    count += 1
        
        # Check if we've reached the Answer Key section
        if current and current.name == 'h3' and 'Answer Key' in current.get_text():
            break
            
        current = current.find_next_sibling()
    
    return explanations



def process_cell_with_images(worksheet, row_idx, col_idx, text, temp_dir):
    """Process a cell that may contain formula images"""
    # Simply set the text with formula URLs in the cell
    cell = worksheet.cell(row=row_idx, column=col_idx)
    cell.value = text

def extract_questions_from_html(html_file_path):
    """Extract questions and options from HTML file"""
    
    # Read the HTML file
    with open(html_file_path, 'r', encoding='utf-8') as file:
        html_content = file.read()
    
    # Parse HTML
    soup = BeautifulSoup(html_content, 'html.parser')
    
    # Extract answer key and explanations first
    answer_key = extract_answer_key(soup)
    explanations = extract_explanations(soup)
    
    # Find all potential question blocks
    all_elements = soup.find_all('p', class_='pull-left clearfix')
    
    questions_data = []
    serial_no = 1
    
    for block in all_elements:
        # Check if we've hit the Explanation or Answer Key section
        # Look for the nearest previous h3 header
        prev_h3 = block.find_previous('h3')
        if prev_h3 and (re.search(r'Explanation', prev_h3.get_text()) or 
                        re.search(r'Answer Key', prev_h3.get_text())):
            # We've reached the Explanation or Answer Key section, stop processing
            break
        # Check if this is a question block (starts with Q.)
        block_text = block.get_text()
        
        if not block_text.strip().startswith('Q.'):
            continue
        
        # Extract question number and text with formulas
        question_text = extract_formula_from_element(block)
        
        # Remove the question number (Q.1), Q.2), etc.) from the beginning
        question_text = re.sub(r'^Q\.\d+\)\s*', '', question_text)
        
        # Find the next sibling elements to get options
        options_dict = {}
        current = block.find_next_sibling()
        
        while current and current.name == 'p' and 'answer' in current.get('class', []):
            # Find the option element within the answer block
            option_element = current.find('p', class_='option pull-left')
            if option_element:
                # Get the full text content with formulas
                full_text = extract_formula_from_element(option_element)
                
                # Extract option letter (a), b), c), etc.) from the beginning
                match = re.match(r'^([a-z])\)', full_text.strip())
                if match:
                    option_letter = match.group(1).upper()
                    # Remove the option letter from the text to get the option content
                    option_text = full_text[len(match.group(0)):].strip()
                    
                    # If option text is empty, try to get text from nested p tags with formulas
                    if not option_text:
                        nested_p = option_element.find('p')
                        if nested_p:
                            option_text = extract_formula_from_element(nested_p)
                    
                    if option_text:
                        options_dict[option_letter] = option_text
            
            current = current.find_next_sibling()
        
        # Create the data entry with empty Subject, Topic, Subtopic
        data_entry = {
            'S No': serial_no,
            'Subject': '',
            'Topic': '',
            'Subtopic': '',
            'Question with Statements': question_text
        }
        
        # Add options to the data entry
        for option_letter, option_text in options_dict.items():
            data_entry[f'Option {option_letter}'] = option_text
        
        # Add answer and explanation
        answer = answer_key.get(serial_no, '')
        explanation = explanations.get(serial_no, '')
        
        
        data_entry['Answer'] = answer
        data_entry['Explanation'] = explanation
        
        questions_data.append(data_entry)
        serial_no += 1
    
    return questions_data

def save_to_excel(questions_data, output_file):
    """Save questions data to Excel file with embedded formula images"""
    # Find all unique option columns across all questions
    all_option_columns = set()
    for data in questions_data:
        option_columns = [col for col in data.keys() if col.startswith('Option ')]
        all_option_columns.update(option_columns)
    
    # Sort option columns alphabetically (Option A, Option B, etc.)
    sorted_option_columns = sorted(list(all_option_columns))
    
    # Define the column order - Answer and Explanation come after option columns
    column_order = ['S No', 'Subject', 'Topic', 'Subtopic', 'Question with Statements'] + sorted_option_columns + ['Answer', 'Explanation']
    
    # Create DataFrame with specified column order
    df = pd.DataFrame(questions_data)
    
    # Ensure all columns exist in the DataFrame
    for col in column_order:
        if col not in df.columns:
            df[col] = ''
    
    # Reorder columns
    df = df[column_order]
    
    # Save to Excel with proper formatting
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Write the dataframe to Excel
        df.to_excel(writer, sheet_name='Questions', index=False)
        
        # Get the worksheet
        worksheet = writer.sheets['Questions']
        
        # Set column widths
        worksheet.column_dimensions['A'].width = 8   # S No
        worksheet.column_dimensions['B'].width = 15  # Subject
        worksheet.column_dimensions['C'].width = 15  # Topic
        worksheet.column_dimensions['D'].width = 15  # Subtopic
        worksheet.column_dimensions['E'].width = 60  # Question with Statements
        
        # Set width for option columns (F onwards)
        from openpyxl.utils import get_column_letter
        for i, option_col in enumerate(sorted_option_columns):
            col_letter = get_column_letter(6 + i)  # F is column 6
            worksheet.column_dimensions[col_letter].width = 30
        
        # Set width for Answer and Explanation columns
        answer_col = get_column_letter(6 + len(sorted_option_columns))  # After all options
        explanation_col = get_column_letter(7 + len(sorted_option_columns))  # After answer
        worksheet.column_dimensions[answer_col].width = 10  # Answer column
        worksheet.column_dimensions[explanation_col].width = 80  # Explanation column
        
        # Process cells with formula images
        for row_idx, row_data in enumerate(questions_data, start=2):  # Start from row 2 (after header)
            for col_idx, col_name in enumerate(column_order, start=1):
                if col_name in row_data:
                    cell_value = str(row_data[col_name])
                    process_cell_with_images(worksheet, row_idx, col_idx, cell_value, None)
        
        # Wrap text for better readability
        from openpyxl.styles import Alignment
        for row in worksheet.iter_rows(min_row=2):
            # Wrap text for all cells
            for cell in row:
                if cell.value:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
        
        # Adjust row heights to accommodate images
        for row in worksheet.iter_rows(min_row=2):
            worksheet.row_dimensions[row[0].row].height = 40  # Set minimum row height
    
    print(f"Excel file saved: {output_file}")

def main():
    # Input and output file paths
    input_html = 'input/Page_618.html'
    output_excel = 'output/questions_output.xlsx'
    
    # Check if input file exists
    if not os.path.exists(input_html):
        print(f"Error: Input file not found: {input_html}")
        sys.exit(1)
    
    print(f"Processing HTML file: {input_html}")
    
    # Extract questions
    questions_data = extract_questions_from_html(input_html)
    
    print(f"Found {len(questions_data)} questions")
    
    # Save to Excel
    save_to_excel(questions_data, output_excel)
    
    # Display first few questions as sample
    if questions_data:
        print("\nSample of extracted questions:")
        for i, q in enumerate(questions_data[:3]):
            print(f"\nS No: {q['S No']}")
            print(f"Question: {q['Question with Statements'][:100]}...")
            # Display options
            option_keys = [k for k in sorted(q.keys()) if k.startswith('Option ')]
            if option_keys:
                print("Options:")
                for opt_key in option_keys:
                    if q.get(opt_key):
                        print(f"  {opt_key}: {q[opt_key][:50]}...")
            # Display answer and explanation
            if q.get('Answer'):
                print(f"Answer: {q['Answer']}")
            if q.get('Explanation'):
                print(f"Explanation: {q['Explanation'][:100]}...")

if __name__ == "__main__":
    main()